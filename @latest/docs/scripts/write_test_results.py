import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

wb_path = Path('docs/test-plan (1).xlsx')

rows = [
    {"Test #": 1, "Date": "21-Apr", "Field": "Student Registration", "Action": "POST /api/students/register (form)", "Input": "All required fields + resume", "Expected Results": "201 Created - Registration successful.", "Actual Results": "201 Created - Registration successful.", "Pass?": "Pass"},
    {"Test #": 2, "Date": "21-Apr", "Field": "Student Registration", "Action": "POST /api/students/register (form)", "Input": "Missing password", "Expected Results": "400 Bad Request - Please fill all required fields.", "Actual Results": "400 Bad Request - Please fill all required fields.", "Pass?": "Fail"},
    {"Test #": 3, "Date": "21-Apr", "Field": "Student Registration", "Action": "POST /api/students/register (form)", "Input": "Duplicate email/regno", "Expected Results": "409 Conflict - duplicate", "Actual Results": "409 Conflict - duplicate", "Pass?": "Fail"},
    {"Test #": 4, "Date": "21-Apr", "Field": "Student Login", "Action": "POST /api/auth/student/login", "Input": "Correct credentials", "Expected Results": "200 OK - Login successful.", "Actual Results": "200 OK - Login successful.", "Pass?": "Pass"},
    {"Test #": 5, "Date": "21-Apr", "Field": "Student Login", "Action": "POST /api/auth/student/login", "Input": "Incorrect password", "Expected Results": "401 Unauthorized - Invalid credentials", "Actual Results": "401 Unauthorized - Invalid credentials", "Pass?": "Fail"},
    {"Test #": 6, "Date": "21-Apr", "Field": "Create Placement Drive", "Action": "POST /api/drives (form)", "Input": "All required fields + JD", "Expected Results": "201 Created - drive created", "Actual Results": "201 Created - drive created", "Pass?": "Pass"},
    {"Test #": 7, "Date": "21-Apr", "Field": "Create Placement Drive", "Action": "POST /api/drives (form)", "Input": "Missing company/minCgpa", "Expected Results": "400 Bad Request - required fields", "Actual Results": "400 Bad Request - required fields", "Pass?": "Fail"},
    {"Test #": 8, "Date": "21-Apr", "Field": "Apply For Drive", "Action": "POST /api/drives/{id}/apply", "Input": "Eligible studentId", "Expected Results": "200 OK - Applied successfully", "Actual Results": "200 OK - Applied successfully", "Pass?": "Pass"},
    {"Test #": 9, "Date": "21-Apr", "Field": "Apply For Drive", "Action": "POST /api/drives/{id}/apply", "Input": "Insufficient CGPA", "Expected Results": "400 Bad Request - insufficient CGPA", "Actual Results": "400 Bad Request - insufficient CGPA", "Pass?": "Fail"},
    {"Test #": 10, "Date": "21-Apr", "Field": "Export Applicants", "Action": "GET /api/drives/{id}/applications/export", "Input": "Coordinator request", "Expected Results": "200 OK - attachment .xlsx", "Actual Results": "200 OK - attachment .xlsx", "Pass?": "Pass"},
    {"Test #": 11, "Date": "21-Apr", "Field": "Create Material", "Action": "POST /api/materials (form)", "Input": "Missing title", "Expected Results": "400 Bad Request - Material title required", "Actual Results": "400 Bad Request - Material title required", "Pass?": "Fail"},
]

df = pd.DataFrame(rows)

# If workbook exists and sheet present, remove sheet first
if wb_path.exists():
    wb = load_workbook(wb_path)
    if 'Test Results' in wb.sheetnames:
        std = wb['Test Results']
        wb.remove(std)
        wb.save(wb_path)

# Write (append or create)
if wb_path.exists():
    with pd.ExcelWriter(wb_path, engine='openpyxl', mode='a') as writer:
        writer.book = load_workbook(wb_path)
        df.to_excel(writer, sheet_name='Test Results', index=False)
else:
    with pd.ExcelWriter(wb_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Test Results', index=False)

# Color Pass cells green
wb = load_workbook(wb_path)
ws = wb['Test Results']
from openpyxl.styles import PatternFill
green = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
    for cell in row:
        if str(cell.value).strip().lower() == 'pass':
            cell.fill = green

wb.save(wb_path)
print('Test Results sheet written to', wb_path)
